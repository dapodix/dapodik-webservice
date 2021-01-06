import os.path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Tuple

from dapodik_webservice import Sekolah
from dapodik_webservice.template import TEMPLATE

MAPPING = {
    "sekolah_id": "C2",
    "nama": "B2",
    "nss": "B3",
    "npsn": "B4",
    "bentuk_pendidikan_id": "B5",
    "bentuk_pendidikan_id_str": "C5",
    "status_sekolah": "B6",
    "status_sekolah_str": "C6",
    "alamat_jalan": "B7",
    "rt": "B8",
    "rw": "B9",
    "kode_wilayah": "B10",
    "kode_pos": "B11",
    "nomor_telepon": "B12",
    "nomor_fax": "B13",
    "email": "B14",
    "website": "B15",
    "is_sks": "B16",
    "lintang": "B17",
    "bujur": "B18",
    "dusun": "B19",
    "desa_kelurahan": "B20",
    "kecamatan": "B21",
    "kabupaten_kota": "B22",
    "provinsi": "B23",
}


def sekolah_export(
    sekolah: Sekolah,
    filepath: str,
    ws: Workbook = None,
    wb: Workbook = None,
    mapping: Dict[str, str] = None,
) -> Tuple[Workbook, Worksheet]:
    if not ws:
        if not wb:
            if os.path.isfile(filepath):
                wb = load_workbook(filepath)
            else:
                wb = load_workbook(TEMPLATE)
        ws = wb["Sekolah"]
    mapping = mapping or MAPPING
    for key, value in MAPPING.items():
        ws[value] = getattr(sekolah, key, ws[value])
    return wb, ws
