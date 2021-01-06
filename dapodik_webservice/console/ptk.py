import os.path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Tuple

from dapodik_webservice import Ptk
from dapodik_webservice.template import TEMPLATE

MAPPING = {
    "tahun_ajaran_id": "B",
    "ptk_terdaftar_id": "C",
    "ptk_id": "D",
    "ptk_induk": "E",
    "tanggal_surat_tugas": "F",
    "nama": "G",
    "jenis_kelamin": "H",
    "tempat_lahir": "I",
    "tanggal_lahir": "J",
    "agama_id": "K",
    "agama_id_str": "L",
    "nuptk": "M",
    "nik": "N",
    "jenis_ptk_id": "O",
    "jenis_ptk_id_str": "P",
    "status_kepegawaian_id": "Q",
    "status_kepegawaian_id_str": "R",
    "nip": "S",
    "pendidikan_terakhir": "T",
    "bidang_studi_terakhir": "U",
    "pangkat_golongan_terakhir": "V",
    # "rwy_pend_formal": "W",
    # "rwy_kepangkatan": "X",
}


def ptk_export(
    list_ptk: List[Ptk],
    filepath: str,
    ws: Worksheet = None,
    wb: Workbook = None,
    start: int = 1,
    table: bool = True,
    mapping: Dict[str, str] = None,
) -> Tuple[Workbook, Worksheet]:
    if not ws:
        if not wb:
            if os.path.isfile(filepath):
                wb = load_workbook(filepath)
            else:
                wb = load_workbook(TEMPLATE)
        ws = wb["Ptk"]
    mapping = mapping or MAPPING
    for index, ptk in enumerate(list_ptk, start + 1):
        ws[f"A{index}"] = index - 1
        for key, value in MAPPING.items():
            cell = f"{value}{index}"
            ws[cell] = getattr(ptk, key, ws[cell])
        end = cell
    if table:
        tab = Table(displayName="PesertaDidik", ref=f"A1:{end}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        ws.add_table(tab)
    return wb, ws
