import os.path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Tuple

from dapodik_webservice import RombonganBelajar
from dapodik_webservice.template import TEMPLATE

MAPPING = {
    "rombongan_belajar_id": "B",
    "nama": "C",
    "tingkat_pendidikan_id": "D",
    "semester_id": "E",
    "jenis_rombel": "F",
    "kurikulum_id": "G",
    "kurikulum_id_str": "H",
    "id_ruang": "I",
    "id_ruang_str": "J",
    "moving_class": "K",
    "ptk_id": "L",
    "ptk_id_str": "M",
    "jenis_rombel_str": "N",
}


def rombongan_belajar_export(
    list_rombongan_belajar: List[RombonganBelajar],
    filepath: str,
    ws: Worksheet = None,
    wb: Workbook = None,
    start: int = 1,
    table: bool = True,
    mapping: Dict[str, str] = None,
) -> Tuple[Workbook, Worksheet]:
    if not ws:
        if not wb:
            if os.path.isfile(filepath):
                wb = load_workbook(filepath)
            else:
                wb = load_workbook(TEMPLATE)
        ws = wb["Rombongan Belajar"]
    mapping = mapping or MAPPING
    for index, rombongan_belajar in enumerate(list_rombongan_belajar, start + 1):
        ws[f"A{index}"] = index - 1
        for key, value in mapping.items():
            cell = f"{value}{index}"
            ws[cell] = getattr(rombongan_belajar, key, ws[cell])
        end = cell
    if table:
        tab = Table(displayName="PesertaDidik", ref=f"A1:{end}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        ws.add_table(tab)
    return wb, ws
