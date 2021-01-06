import os.path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Tuple

from dapodik_webservice import PesertaDidik
from dapodik_webservice.template import TEMPLATE

MAPPING = {
    "registrasi_id": "B",
    "jenis_pendaftaran_id": "C",
    "jenis_pendaftaran_id_str": "D",
    "nipd": "E",
    "tanggal_masuk_sekolah": "F",
    "sekolah_asal": "G",
    "peserta_didik_id": "H",
    "nama": "I",
    "nisn": "J",
    "jenis_kelamin": "K",
    "nik": "L",
    "tempat_lahir": "M",
    "tanggal_lahir": "N",
    "agama_id": "O",
    "agama_id_str": "P",
    "alamat_jalan": "Q",
    "nomor_telepon_rumah": "R",
    "nomor_telepon_seluler": "S",
    "nama_ayah": "T",
    "pekerjaan_ayah_id": "U",
    "pekerjaan_ayah_id_str": "V",
    "nama_ibu": "W",
    "pekerjaan_ibu_id": "X",
    "pekerjaan_ibu_id_str": "Y",
    "nama_wali": "Z",
    "pekerjaan_wali_id": "AA",
    "pekerjaan_wali_id_str": "AB",
    "tinggi_badan": "AC",
    "berat_badan": "AD",
    "semester_id": "AE",
    "anggota_rombel_id": "AF",
    "rombongan_belajar_id": "AG",
    "tingkat_pendidikan_id": "AH",
    "nama_rombel": "AI",
    "kurikulum_id": "AJ",
    "kurikulum_id_str": "AK",
}


def peserta_didik_export(
    list_peserta_didik: List[PesertaDidik],
    filepath: str,
    ws: Worksheet = None,
    wb: Workbook = None,
    start: int = 1,
    table: bool = True,
    mapping: Dict[str, str] = None,
) -> Tuple[Workbook, Worksheet]:
    if not ws:
        if not wb:
            if os.path.isfile(filepath):
                wb = load_workbook(filepath)
            else:
                wb = load_workbook(TEMPLATE)
        ws = wb["Peserta Didik"]
    mapping = mapping or MAPPING
    for index, peserta_didik in enumerate(list_peserta_didik, start + 1):
        ws[f"A{index}"] = index - 1
        for key, value in mapping.items():
            cell = f"{value}{index}"
            ws[cell] = getattr(peserta_didik, key, ws[cell])
        end = cell
    if table:
        tab = Table(displayName="PesertaDidik", ref=f"A1:{end}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        ws.add_table(tab)
    return wb, ws
