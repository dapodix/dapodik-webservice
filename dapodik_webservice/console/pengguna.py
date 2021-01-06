import os.path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Tuple

from dapodik_webservice import Pengguna
from dapodik_webservice.template import TEMPLATE

MAPPING = {
    "pengguna_id": "B",
    "sekolah_id": "C",
    "username": "D",
    "nama": "E",
    "peran_id_str": "F",
    "password": "G",
    "alamat": "H",
    "no_telepon": "I",
    "no_hp": "J",
    "ptk_id": "K",
    "peserta_didik_id": "L",
}


def pengguna_export(
    list_pengguna: List[Pengguna],
    filepath: str,
    ws: Worksheet = None,
    wb: Workbook = None,
    start: int = 1,
    table: bool = True,
    mapping: Dict[str, str] = None,
) -> Tuple[Workbook, Worksheet]:
    if not ws:
        if not wb:
            if os.path.isfile(filepath):
                wb = load_workbook(filepath)
            else:
                wb = load_workbook(TEMPLATE)
        ws = wb["Pengguna"]
    mapping = mapping or MAPPING
    end = "A1"
    for index, pengguna in enumerate(list_pengguna, start + 1):
        ws[f"A{index}"] = index - 1
        for properti, huruf in mapping.items():
            cell = f"{huruf}{index}"
            ws[cell] = getattr(pengguna, properti, ws[cell])
        end = cell
    if table:
        tab = Table(displayName="Pengguna", ref=f"A1:{end}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        ws.add_table(tab)
    return wb, ws
