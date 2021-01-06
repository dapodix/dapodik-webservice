import os
from openpyxl import load_workbook, Workbook

CWD = os.getcwd()


def makefilepath(filename: str) -> str:
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    return os.path.join(CWD, filename)


def get_workbook(filepath: str) -> Workbook:
    if os.path.isfile(filepath):
        return load_workbook(filepath)
    return Workbook()
